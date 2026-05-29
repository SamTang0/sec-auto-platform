from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .logger import get_logger

logger = get_logger("excel")

class ExcelExporter:
    def __init__(self, output_dir: Path = None):
        self.output_dir = output_dir or Path(__file__).parent.parent / "reports"
        self.output_dir.mkdir(exist_ok=True)

    def export_alerts(self, alerts: list, filename: str = None) -> Path:
        if not alerts:
            return None
        if not filename:
            filename = f"alerts_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = self.output_dir / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "告警列表"
        headers = ["告警ID", "标题", "风险等级", "Orca评分", "状态", "创建时间"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for r, a in enumerate(alerts, 2):
            ws.cell(r, 1, a.get('AlertId', ''))
            ws.cell(r, 2, a.get('Title', '')[:60])
            ws.cell(r, 3, a.get('RiskLevel', ''))
            ws.cell(r, 4, a.get('OrcaScore', ''))
            ws.cell(r, 5, a.get('Status', ''))
            ws.cell(r, 6, a.get('CreatedAt', ''))
        wb.save(path)
        logger.info(f"导出告警: {path}")
        return path

    def export_assets(self, assets: list, filename: str = None) -> Path:
        if not assets:
            return None
        if not filename:
            filename = f"assets_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = self.output_dir / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "资产列表"
        headers = ["资产名称", "类型", "风险等级", "Orca评分", "公网暴露", "首次发现"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for r, a in enumerate(assets, 2):
            ws.cell(r, 1, a.get('Name', ''))
            ws.cell(r, 2, a.get('Type', ''))
            ws.cell(r, 3, a.get('RiskLevel', ''))
            ws.cell(r, 4, a.get('OrcaScore', ''))
            ws.cell(r, 5, '是' if a.get('IsInternetFacing') else '否')
            ws.cell(r, 6, a.get('FirstSeen', ''))
        wb.save(path)
        logger.info(f"导出资产: {path}")
        return path
